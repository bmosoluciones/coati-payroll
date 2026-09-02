# SPDX-License-Identifier: Apache-2.0
# SPDX-FileCopyrightText: 2025 - 2026 BMO Soluciones, S.A.
"""Initial benefit balance loading views."""

from __future__ import annotations

from decimal import Decimal
from datetime import datetime

from flask import Blueprint, abort, flash, redirect, render_template, request, url_for, Response
from flask_login import current_user
from sqlalchemy import and_

from coati_payroll.enums import CargaInicialEstado
from coati_payroll.forms import CargaInicialPrestacionForm
from coati_payroll.i18n import _
from coati_payroll.model import (
    CargaInicialPrestacion,
    Empleado,
    Moneda,
    Prestacion,
    PrestacionAcumulada,
    empresa_prestacion,
    db,
)
from coati_payroll.rbac import require_read_access, require_write_access
from coati_payroll.tenant import (
    concept_scope_query,
    require_company_access,
    scope_company_query,
    scoped_employee_owned_or_404,
)
from coati_payroll.vistas.constants import PER_PAGE

carga_inicial_prestacion_bp = Blueprint("carga_inicial_prestacion", __name__, url_prefix="/carga-inicial-prestaciones")
INITIAL_BALANCE_INDEX_ENDPOINT = "carga_inicial_prestacion.index"
INITIAL_BALANCE_FORM_TEMPLATE = "modules/carga_inicial_prestacion/form.html"


@carga_inicial_prestacion_bp.route("/", methods=["GET"])
@require_read_access()
def index():
    """List all initial benefit balance loads."""
    page = request.args.get("page", 1, type=int)
    estado_filter = request.args.get("estado", "")

    query = scope_company_query(
        db.select(CargaInicialPrestacion).join(CargaInicialPrestacion.empleado),
        Empleado.empresa_id,
    )

    # Apply filters
    if estado_filter:
        query = query.filter(CargaInicialPrestacion.estado == estado_filter)

    # Order by creation date descending
    query = query.order_by(CargaInicialPrestacion.creado.desc())

    # Paginate
    pagination = db.paginate(query, page=page, per_page=PER_PAGE, error_out=False)
    cargas = pagination.items

    return render_template(
        "modules/carga_inicial_prestacion/index.html",
        cargas=cargas,
        pagination=pagination,
        estado_filter=estado_filter,
    )


@carga_inicial_prestacion_bp.route("/nueva", methods=["GET", "POST"])
@require_write_access()
def nueva():
    """Create a new initial benefit balance load."""
    form = CargaInicialPrestacionForm()

    # Populate select field choices
    form.empleado_id.choices = [("", _("-- Seleccionar --"))] + [
        (emp.id, f"{emp.codigo_empleado} - {emp.primer_nombre} {emp.primer_apellido}")
        for emp in db.session.execute(
            scope_company_query(db.select(Empleado), Empleado.empresa_id)
            .filter(Empleado.activo.is_(True))
            .order_by(Empleado.codigo_empleado)
        )
        .scalars()
        .all()
    ]

    form.prestacion_id.choices = [("", _("-- Seleccionar --"))] + [
        (prest.id, f"{prest.codigo} - {prest.nombre}")
        for prest in db.session.execute(
            concept_scope_query(db.select(Prestacion), empresa_prestacion, Prestacion.id)
            .filter(Prestacion.activo.is_(True))
            .order_by(Prestacion.codigo)
        )
        .scalars()
        .all()
    ]

    form.moneda_id.choices = [("", _("-- Seleccionar --"))] + [
        (mon.id, f"{mon.codigo} - {mon.nombre}")
        for mon in Moneda.query.filter_by(activo=True).order_by(Moneda.codigo).all()
    ]

    if form.validate_on_submit():
        empleado = db.session.get(Empleado, form.empleado_id.data)
        if empleado is None:
            abort(404)
        require_company_access(empleado.empresa_id)

        # Check for duplicate
        existing_query = scope_company_query(
            db.select(CargaInicialPrestacion).join(CargaInicialPrestacion.empleado),
            Empleado.empresa_id,
        ).filter(
            and_(
                CargaInicialPrestacion.empleado_id == form.empleado_id.data,
                CargaInicialPrestacion.prestacion_id == form.prestacion_id.data,
                CargaInicialPrestacion.anio_corte == form.anio_corte.data,
                CargaInicialPrestacion.mes_corte == form.mes_corte.data,
            )
        )
        existing = db.session.execute(existing_query).scalar_one_or_none()

        if existing:
            flash(
                _("Ya existe una carga inicial para este empleado, prestación y periodo."),
                "warning",
            )
            return render_template(INITIAL_BALANCE_FORM_TEMPLATE, form=form)

        carga = CargaInicialPrestacion(
            empleado_id=form.empleado_id.data,
            prestacion_id=form.prestacion_id.data,
            anio_corte=form.anio_corte.data,
            mes_corte=form.mes_corte.data,
            moneda_id=form.moneda_id.data,
            saldo_acumulado=form.saldo_acumulado.data if form.saldo_acumulado.data is not None else Decimal("0.00"),
            tipo_cambio=form.tipo_cambio.data if form.tipo_cambio.data is not None else Decimal("1.0"),
            saldo_convertido=form.saldo_convertido.data if form.saldo_convertido.data is not None else Decimal("0.00"),
            observaciones=form.observaciones.data,
            estado="draft",
            creado_por=current_user.usuario if current_user.is_authenticated else None,
        )

        db.session.add(carga)
        db.session.commit()

        flash(_("Carga inicial creada exitosamente en estado borrador."), "success")
        return redirect(url_for(INITIAL_BALANCE_INDEX_ENDPOINT))

    return render_template(INITIAL_BALANCE_FORM_TEMPLATE, form=form)


@carga_inicial_prestacion_bp.route("/<carga_id>/editar", methods=["GET", "POST"])
@require_write_access()
def editar(carga_id):
    """Edit an initial benefit balance load (only if in draft status)."""
    carga = scoped_employee_owned_or_404(CargaInicialPrestacion, carga_id, CargaInicialPrestacion.empleado_id)

    if carga.estado == "applied":
        flash(_("No se puede editar una carga inicial ya aplicada."), "warning")
        return redirect(url_for(INITIAL_BALANCE_INDEX_ENDPOINT))

    form = CargaInicialPrestacionForm(obj=carga)

    # Populate select field choices
    form.empleado_id.choices = [("", _("-- Seleccionar --"))] + [
        (emp.id, f"{emp.codigo_empleado} - {emp.primer_nombre} {emp.primer_apellido}")
        for emp in db.session.execute(
            scope_company_query(db.select(Empleado), Empleado.empresa_id)
            .filter(Empleado.activo.is_(True))
            .order_by(Empleado.codigo_empleado)
        )
        .scalars()
        .all()
    ]

    form.prestacion_id.choices = [("", _("-- Seleccionar --"))] + [
        (prest.id, f"{prest.codigo} - {prest.nombre}")
        for prest in db.session.execute(
            concept_scope_query(db.select(Prestacion), empresa_prestacion, Prestacion.id)
            .filter(Prestacion.activo.is_(True))
            .order_by(Prestacion.codigo)
        )
        .scalars()
        .all()
    ]

    form.moneda_id.choices = [("", _("-- Seleccionar --"))] + [
        (mon.id, f"{mon.codigo} - {mon.nombre}")
        for mon in Moneda.query.filter_by(activo=True).order_by(Moneda.codigo).all()
    ]

    if form.validate_on_submit():
        empleado = db.session.get(Empleado, form.empleado_id.data)
        if empleado is None:
            abort(404)
        require_company_access(empleado.empresa_id)
        carga.empleado_id = form.empleado_id.data
        carga.prestacion_id = form.prestacion_id.data
        carga.anio_corte = form.anio_corte.data
        carga.mes_corte = form.mes_corte.data
        carga.moneda_id = form.moneda_id.data
        carga.saldo_acumulado = form.saldo_acumulado.data if form.saldo_acumulado.data is not None else Decimal("0.00")
        carga.tipo_cambio = form.tipo_cambio.data if form.tipo_cambio.data is not None else Decimal("1.0")
        carga.saldo_convertido = (
            form.saldo_convertido.data if form.saldo_convertido.data is not None else Decimal("0.00")
        )
        carga.observaciones = form.observaciones.data
        carga.modificado_por = current_user.usuario if current_user.is_authenticated else None

        db.session.commit()

        flash(_("Carga inicial actualizada exitosamente."), "success")
        return redirect(url_for(INITIAL_BALANCE_INDEX_ENDPOINT))

    return render_template(INITIAL_BALANCE_FORM_TEMPLATE, form=form, carga=carga)


@carga_inicial_prestacion_bp.route("/<carga_id>/aplicar", methods=["POST"])
@require_write_access()
def aplicar(carga_id):
    """Apply an initial balance load - creates transaction in prestacion_acumulada."""
    carga = scoped_employee_owned_or_404(CargaInicialPrestacion, carga_id, CargaInicialPrestacion.empleado_id)

    if carga.estado == CargaInicialEstado.APLICADO:
        flash(_("Esta carga inicial ya ha sido aplicada."), "warning")
        return redirect(url_for(INITIAL_BALANCE_INDEX_ENDPOINT))

    try:
        # Create transaction in prestacion_acumulada
        transaccion = PrestacionAcumulada(
            empleado_id=carga.empleado_id,
            prestacion_id=carga.prestacion_id,
            fecha_transaccion=datetime.now().date(),
            tipo_transaccion="saldo_inicial",
            anio=carga.anio_corte,
            mes=carga.mes_corte,
            moneda_id=carga.moneda_id,
            monto_transaccion=carga.saldo_convertido,
            saldo_anterior=Decimal("0.00"),
            saldo_nuevo=carga.saldo_convertido,
            carga_inicial_id=carga.id,
            # Payroll provisions read their prior balance within the employee's
            # company.  Keep the initial balance in that same ledger scope.
            empresa_id=carga.empleado.empresa_id,
            observaciones=f"Carga inicial - {carga.observaciones or ''}",
            procesado_por=current_user.usuario if current_user.is_authenticated else None,
            creado_por=current_user.usuario if current_user.is_authenticated else None,
        )

        db.session.add(transaccion)

        # Update carga status
        carga.estado = CargaInicialEstado.APLICADO
        carga.fecha_aplicacion = datetime.now()
        carga.aplicado_por = current_user.usuario if current_user.is_authenticated else None
        carga.modificado_por = current_user.usuario if current_user.is_authenticated else None

        db.session.commit()

        flash(_("Carga inicial aplicada exitosamente."), "success")

    except Exception as e:
        db.session.rollback()
        flash(_("Error al aplicar la carga inicial: %(error)s", error=str(e)), "danger")

    return redirect(url_for(INITIAL_BALANCE_INDEX_ENDPOINT))


@carga_inicial_prestacion_bp.route("/<carga_id>/eliminar", methods=["POST"])
@require_write_access()
def eliminar(carga_id):
    """Delete an initial balance load (only if in draft status)."""
    carga = scoped_employee_owned_or_404(CargaInicialPrestacion, carga_id, CargaInicialPrestacion.empleado_id)

    if carga.estado == CargaInicialEstado.APLICADO:
        flash(_("No se puede eliminar una carga inicial ya aplicada."), "warning")
        return redirect(url_for(INITIAL_BALANCE_INDEX_ENDPOINT))

    try:
        db.session.delete(carga)
        db.session.commit()
        flash(_("Carga inicial eliminada exitosamente."), "success")
    except Exception as e:
        db.session.rollback()
        flash(_("Error al eliminar la carga inicial: %(error)s", error=str(e)), "danger")

    return redirect(url_for(INITIAL_BALANCE_INDEX_ENDPOINT))


@carga_inicial_prestacion_bp.route("/reporte", methods=["GET"])
@require_read_access()
def reporte():
    """Generate accumulated benefits report."""
    # Get filter parameters
    empleado_id = request.args.get("empleado_id")
    prestacion_id = request.args.get("prestacion_id")
    fecha_desde = request.args.get("fecha_desde")
    fecha_hasta = request.args.get("fecha_hasta")

    # Build query
    query = scope_company_query(
        db.select(PrestacionAcumulada).join(PrestacionAcumulada.empleado),
        Empleado.empresa_id,
    )

    if empleado_id:
        query = query.filter(PrestacionAcumulada.empleado_id == empleado_id)

    if prestacion_id:
        query = query.filter(PrestacionAcumulada.prestacion_id == prestacion_id)

    if fecha_desde:
        query = query.filter(PrestacionAcumulada.fecha_transaccion >= fecha_desde)

    if fecha_hasta:
        query = query.filter(PrestacionAcumulada.fecha_transaccion <= fecha_hasta)

    # Order by date
    transacciones = (
        db.session.execute(
            query.order_by(
                PrestacionAcumulada.empleado_id,
                PrestacionAcumulada.prestacion_id,
                PrestacionAcumulada.fecha_transaccion,
            )
        )
        .scalars()
        .all()
    )

    # Get choices for filters
    empleados = (
        db.session.execute(
            scope_company_query(db.select(Empleado), Empleado.empresa_id)
            .filter(Empleado.activo.is_(True))
            .order_by(Empleado.codigo_empleado)
        )
        .scalars()
        .all()
    )
    prestaciones = (
        db.session.execute(
            concept_scope_query(db.select(Prestacion), empresa_prestacion, Prestacion.id)
            .filter(Prestacion.activo.is_(True))
            .order_by(Prestacion.codigo)
        )
        .scalars()
        .all()
    )

    return render_template(
        "modules/carga_inicial_prestacion/reporte.html",
        transacciones=transacciones,
        empleados=empleados,
        prestaciones=prestaciones,
        empleado_id=empleado_id,
        prestacion_id=prestacion_id,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
    )


def _get_report_transactions():
    """Load report transactions using the optional request filters."""
    query = scope_company_query(
        db.select(PrestacionAcumulada).join(PrestacionAcumulada.empleado),
        Empleado.empresa_id,
    )
    filters = (
        ("empleado_id", PrestacionAcumulada.empleado_id),
        ("prestacion_id", PrestacionAcumulada.prestacion_id),
        ("fecha_desde", PrestacionAcumulada.fecha_transaccion),
        ("fecha_hasta", PrestacionAcumulada.fecha_transaccion),
    )
    operators = (
        lambda field, value: field == value,
        lambda field, value: field == value,
        lambda field, value: field >= value,
        lambda field, value: field <= value,
    )
    for (parameter, field), operator in zip(filters, operators):
        value = request.args.get(parameter)
        if value:
            query = query.filter(operator(field, value))
    query = query.order_by(
        PrestacionAcumulada.empleado_id,
        PrestacionAcumulada.prestacion_id,
        PrestacionAcumulada.fecha_transaccion,
    )
    return db.session.execute(query).scalars().all()


def _report_row_values(trans):
    """Return the audit fields for one accumulated-benefit transaction."""
    return (
        trans.id,
        trans.fecha_transaccion.strftime("%Y-%m-%d"),
        trans.empleado.codigo_empleado,
        f"{trans.empleado.primer_nombre} {trans.empleado.primer_apellido}",
        trans.prestacion.codigo,
        trans.prestacion.nombre,
        trans.prestacion.tipo_acumulacion,
        trans.tipo_transaccion,
        trans.anio,
        trans.mes,
        float(trans.monto_transaccion),
        float(trans.saldo_anterior),
        float(trans.saldo_nuevo),
        trans.moneda.codigo,
        trans.nomina_id or "",
        trans.carga_inicial_id or "",
        trans.procesado_por or "",
        trans.creado.strftime("%Y-%m-%d"),
        trans.creado_por or "",
        trans.observaciones or "",
    )


def _write_report_rows(ws, transactions):
    """Write transaction rows to the report worksheet."""
    for row_num, trans in enumerate(transactions, 2):
        for column_num, value in enumerate(_report_row_values(trans), 1):
            ws.cell(row=row_num, column=column_num, value=value)


def _set_report_column_widths(ws):
    """Set bounded widths based on the rendered worksheet values."""
    for column in ws.columns:
        values = (str(cell.value) for cell in column if cell.value is not None)
        max_length = max((len(value) for value in values), default=0)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)


@carga_inicial_prestacion_bp.route("/reporte/excel", methods=["GET"])
@require_read_access()
def reporte_excel():
    """Export accumulated benefits report to Excel."""
    import io

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        flash(_("La librería openpyxl no está instalada. No se puede generar el reporte Excel."), "danger")
        return redirect(url_for("carga_inicial_prestacion.reporte"))

    transacciones = _get_report_transactions()

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Prestaciones Acumuladas"

    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Headers - Enhanced for audit purposes
    headers = [
        "ID Transacción",
        "Fecha Transacción",
        "Código Empleado",
        "Empleado",
        "Código Prestación",
        "Prestación",
        "Tipo Acumulación",
        "Tipo Transacción",
        "Año",
        "Mes",
        "Monto Transacción",
        "Saldo Anterior",
        "Saldo Nuevo",
        "Moneda",
        "Nómina ID",
        "Carga Inicial ID",
        "Procesado Por",
        "Fecha Creación",
        "Creado Por",
        "Observaciones",
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font

    _write_report_rows(ws, transacciones)

    _set_report_column_widths(ws)

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=prestaciones_acumuladas.xlsx"},
    )
