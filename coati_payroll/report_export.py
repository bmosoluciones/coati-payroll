# SPDX-License-Identifier: Apache-2.0
# SPDX-FileCopyrightText: 2025 - 2026 BMO Soluciones, S.A.
"""Report export functionality for Excel format.

Provides utilities to export report results to Excel files with proper
formatting and metadata.
"""

from __future__ import annotations

# <-------------------------------------------------------------------------> #
# Standard library
# <-------------------------------------------------------------------------> #
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional
from html import escape

# <-------------------------------------------------------------------------> #
# Third party libraries
# <-------------------------------------------------------------------------> #
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from weasyprint import HTML

    WEASYPRINT_AVAILABLE = True
except ImportError:
    WEASYPRINT_AVAILABLE = False


# <-------------------------------------------------------------------------> #
# Local modules
# <-------------------------------------------------------------------------> #
from coati_payroll.config import DIRECTORIO_APP
from coati_payroll.log import log


class ReportExporter:
    """Handles exporting report results to various formats."""

    def __init__(self, report_name: str, results: List[Dict[str, Any]]):
        """Initialize exporter.

        Args:
            report_name: Name of the report
            results: List of result dictionaries
        """
        self.report_name = report_name
        self.results = results

    def to_excel(self, output_path: Optional[str] = None) -> str:
        """Export results to Excel format.

        Args:
            output_path: Optional output file path. If not provided, generates one.

        Returns:
            Path to exported file

        Raises:
            ImportError: If openpyxl is not installed
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl")

        output_path = self._resolve_output_path(output_path, ".xlsx")

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = self.report_name[:31]  # Excel sheet name limit

        self._write_metadata(ws)
        self._write_result_table(ws, start_row=5)

        # Save workbook
        wb.save(output_path)
        log.info("Report exported to: %s", output_path)

        return output_path

    def _resolve_output_path(self, output_path: Optional[str], extension: str) -> str:
        """Return the requested path or create a safe default export path."""
        if output_path:
            return output_path
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = "".join(c for c in self.report_name if c.isalnum() or c in (" ", "_", "-")).strip()
        filename = f"{safe_name.replace(' ', '_')}_{timestamp}{extension}"
        exports_dir = Path(DIRECTORIO_APP) / "exports" / "reports"
        exports_dir.mkdir(parents=True, exist_ok=True)
        return str((exports_dir / filename).absolute())

    def _write_metadata(self, worksheet) -> None:
        """Write report metadata and formatting to a worksheet."""
        metadata = [
            ("A1", "Report:", self.report_name),
            ("A2", "Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("A3", "Total Records:", len(self.results)),
        ]
        for label_cell, label, value in metadata:
            worksheet[label_cell] = label
            worksheet[label_cell].font = Font(bold=True)
            worksheet.cell(row=worksheet[label_cell].row, column=2, value=value)

    def _write_result_table(self, worksheet, start_row: int) -> None:
        """Write result rows and adjust their column widths."""
        if not self.results:
            return
        headers = list(self.results[0].keys())
        self._write_headers(worksheet, headers, start_row)
        for row_idx, row_data in enumerate(self.results, start=start_row + 1):
            for col_idx, header in enumerate(headers, start=1):
                worksheet.cell(row=row_idx, column=col_idx, value=row_data.get(header))
        for col_idx, header in enumerate(headers, start=1):
            values = [len(str(row_data.get(header))) for row_data in self.results if row_data.get(header) is not None]
            max_length = max([len(str(header)), *values])
            worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

    @staticmethod
    def _write_headers(worksheet, headers: list[str], row: int) -> None:
        """Write and style result table headers."""
        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def to_csv(self, output_path: Optional[str] = None) -> str:
        """Export results to CSV format.

        Args:
            output_path: Optional output file path

        Returns:
            Path to exported file
        """
        import csv

        # Generate output path if not provided
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_name = "".join(c for c in self.report_name if c.isalnum() or c in (" ", "_", "-")).strip()
            safe_name = safe_name.replace(" ", "_")  # Replace spaces with underscores for Windows compatibility
            filename = f"{safe_name}_{timestamp}.csv"

            exports_dir = Path(DIRECTORIO_APP) / "exports" / "reports"
            exports_dir.mkdir(parents=True, exist_ok=True)

            output_path = str((exports_dir / filename).absolute())

        # Write CSV
        if self.results:
            headers = list(self.results[0].keys())

            with open(output_path, "w", newline="", encoding="utf-8") as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=headers)
                writer.writeheader()
                writer.writerows(self.results)

        log.info("Report exported to: %s", output_path)
        return output_path

    def to_pdf(self, output_path: Optional[str] = None) -> str:
        """Export the tabular report as a printable PDF."""
        if not WEASYPRINT_AVAILABLE:
            raise ImportError("weasyprint is required for PDF export")
        output_path = self._resolve_output_path(output_path, ".pdf")
        headers = list(self.results[0].keys()) if self.results else []
        header_html = "".join(f"<th>{escape(str(header))}</th>" for header in headers)
        body_html = "".join(
            "<tr>" + "".join(f"<td>{escape(str(row.get(header, '')))}</td>" for header in headers) + "</tr>"
            for row in self.results
        )
        html = f"""<html><head><meta charset='utf-8'><style>
        @page {{ size: A4 landscape; margin: 1cm; }} body {{ font-family: sans-serif; font-size: 9pt; }}
        h1 {{ font-size: 16pt; }} table {{ border-collapse: collapse; width: 100%; }}
        th, td {{ border: 1px solid #999; padding: 4px; text-align: left; }} th {{ background: #366092; color: white; }}
        </style></head><body><h1>{escape(self.report_name)}</h1>
        <p>Generated: {escape(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))}</p>
        <table><thead><tr>{header_html}</tr></thead><tbody>{body_html}</tbody></table>
        </body></html>"""
        HTML(string=html).write_pdf(output_path)
        log.info("Report exported to PDF: %s", output_path)
        return output_path


def export_report_to_excel(report_name: str, results: List[Dict[str, Any]], output_path: Optional[str] = None) -> str:
    """Convenience function to export report to Excel.

    Args:
        report_name: Name of the report
        results: Report results
        output_path: Optional output path

    Returns:
        Path to exported file
    """
    exporter = ReportExporter(report_name, results)
    return exporter.to_excel(output_path)


def export_report_to_csv(report_name: str, results: List[Dict[str, Any]], output_path: Optional[str] = None) -> str:
    """Convenience function to export report to CSV.

    Args:
        report_name: Name of the report
        results: Report results
        output_path: Optional output path

    Returns:
        Path to exported file
    """
    exporter = ReportExporter(report_name, results)
    return exporter.to_csv(output_path)


def export_report_to_pdf(report_name: str, results: List[Dict[str, Any]], output_path: Optional[str] = None) -> str:
    """Convenience function to export a report to PDF."""
    return ReportExporter(report_name, results).to_pdf(output_path)
